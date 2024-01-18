import random
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime, timedelta

class Defense():
    def __init__(self, surname, name, thesis, degree, form, speciality, promoter, reviewer):
        self.surname = surname
        self.name = name
        self.thesis = thesis
        self.degree = degree
        self.form = form
        self.speciality = speciality
        self.promoter = promoter  # Academic class instance
        self.reviewer = reviewer  # Academic class instance

class Slot():
    def __init__(self, date, hour, duration, chairman, hall_no = 1):
        self.date = date
        self.hour = hour
        self.duration = duration
        self.chairman = chairman
        self.hall = hall_no

class DefenseTerm():
    def __init__(self, defense, slot):
        self.defense = defense
        self.slot = slot

    def return_academics(self):
        return [self.slot.chairman, self.defense.promoter, self.defense.reviewer]

    def is_academic_in_defense(self, academic):
        return academic in self.return_academics()

    def check_academics_collisions(self, other):
        collisions = 0
        for i in other.return_academics():
            if self.is_academic_in_defense(i):
                collisions += 1
        return collisions

class DefenseList():
    def __init__(self, filename):
        self.academics, self.defense_list = self.load_defenses_from_file(filename)

    def __getitem__(self, index):
        return self.defense_list[index]

    def load_defenses_from_file(self, filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Extract list of academics
        academics_set = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            academics_set.add(str(row[6]).strip().lower())
            academics_set.add(str(row[7]).strip().lower())

        academics_list = [Academic(academic_name) for academic_name in academics_set]

        defense_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            surname, name, thesis, degree, form, speciality, promoter_name, reviewer_name = map(
                lambda x: str(x).strip(), row[:8])

            promoter = next((academic for academic in academics_list if academic.name.lower() == promoter_name.lower()),
                            None)
            reviewer = next((academic for academic in academics_list if academic.name.lower() == reviewer_name.lower()),
                            None)

            if promoter is None:
                print(f"Promoter not found for {surname}, {name}.")
            if reviewer is None:
                print(f"Reviewer not found for {surname}, {name}.")

            defense = Defense(surname, name, thesis, degree, form, speciality, promoter, reviewer)
            defense_list.append(defense)

        return academics_list, defense_list

class DefensesTermsList():
    def __init__(self, defenses_terms):
        self.defenses_terms = defenses_terms

    def print_list(self):
        for term in self.defenses_terms:
            print(
                f"{term.defense.surname} {term.defense.name} - Date: {term.date}, Hour: {term.hour}, Chairman: {term.chairman}")

    def save_to_xlsx(self, filename):
        # Create a DataFrame with the information about defenses
        data = {
            'Surname': [term.defense.surname for term in self.defenses_terms],
            'Name': [term.defense.name for term in self.defenses_terms],
            'Date': [term.slot.date for term in self.defenses_terms],
            'Hour': [term.slot.hour for term in self.defenses_terms],
            'Chairman': [term.slot.chairman for term in self.defenses_terms],
            'Promoter': [term.defense.promoter.name for term in self.defenses_terms],
            'Reviewer': [term.defense.reviewer.name for term in self.defenses_terms],
            'Hall_no': [term.slot.hall for term in self.defenses_terms]
        }
        df = pd.DataFrame(data)

        # Sort the DataFrame by 'Date' and 'Hour'
        df['Hour'] = pd.to_datetime(df['Hour'], format='%H:%M:%S').dt.time
        df = df.sort_values(by=['Date', 'Hour']).reset_index(drop=True)
        df['Date'] = pd.to_datetime(df['Date'])
        df.to_excel(filename, index=False)

    def mix_value(self):
        mix_value = 0

        # Calculate points for consecutive sequences of the same promoter
        promoter_sequences = self.calculate_consecutive_sequences([term.defense.promoter for term in self.defenses_terms])
        for length in promoter_sequences:
            mix_value += 10 * length - length**2

        # Calculate points for consecutive sequences of the same reviewer
        reviewer_sequences = self.calculate_consecutive_sequences([term.defense.reviewer for term in self.defenses_terms])
        for length in reviewer_sequences:
            mix_value += 10 * length - length**2
        return mix_value

    def calculate_consecutive_sequences(self, academic_list):
        consecutive_sequences = []
        current_length = 1

        for i in range(1, len(academic_list)):
            if academic_list[i] == academic_list[i - 1]:
                current_length += 1
            else:
                consecutive_sequences.append(current_length)
                current_length = 1

        consecutive_sequences.append(current_length)  # Include the last sequence
        return consecutive_sequences

    def correction_score(self):
        points = 0

        for term in self.defenses_terms:
            # Sprawdzenie czy przewodniczący nie zajmuje innej pozycji
            if term.slot.chairman == term.defense.promoter or term.slot.chairman == term.defense.reviewer:
                points -= 1  # Minus za złe przypisanie przewodniczącego
            else:
                points += 1  # Plus za poprawne przypisanie przewodniczącego

            hall_colisions = -1
            person_colisions = -1
            for i in range(len(self.defenses_terms)):
                if term.slot.date == self.defenses_terms[i].slot.date:
                    if term.slot.hour == self.defenses_terms[i].slot.hour:
                        if term.slot.hall == self.defenses_terms[i].slot.hall:
                            hall_colisions += 1
                        person_colisions += term.check_academics_collisions(self.defenses_terms[i])

            # Sprawdzanie kolidujących obron i przyznawanie punktów
            points -= hall_colisions  # Minus za kolidujące sale
            points -= person_colisions  # Minus za kolidujące osoby
        return points

    def wishes_score(self, wish_list):
        conflicts = 0
        for i in self.defenses_terms:
            conflicts += wish_list.count_conflicts(i.slot.date, i.slot.hour, i.slot.hour)#add minutes into end time
        return conflicts

    def fitness(self, wish_list):
        value = 0
        value += self.mix_value()
        value += self.correction_score()
        value -= self.wishes_score(wish_list)
        return value
    def print_conflicts(self):
        print("conflicts evaluation:",self.correction_score())


class SlotsList:
    def __init__(self, filename):
        self.filename = filename
        self.slots = self.load_data_from_file()
        self.slot_duration = timedelta(minutes=30)

    def load_data_from_file(self):
        slots = []
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                date, start_time, duration,chairman, hall_no = row[:5]
                slot = Slot(date, start_time, duration, chairman, hall_no)
                slots.append(slot)

        except Exception as e:
            print(f"Error loading slots from file: {e}")

        return slots

class WishList():
    def __init__(self, filename):
        self.wishlist = self.load_from_file(filename)

    def __getitem__(self, index):
        return self.wishlist[index]

    def load_from_file(self, filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        wishlist = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            academic_name, date, start_time, end_time = row[:4]

            # Create Academic object
            academic = Academic(academic_name)

            # Create Wish object and add it to the wishlist
            wish = Wish(academic, date, start_time, end_time)
            wishlist.append(wish)

        return wishlist

    def count_conflicts(self,date, start_time, end_time):
        counter = 0
        for i in self.wishlist:
            if i.check_availability(date, start_time, end_time):
                counter += 1

        return counter


from datetime import datetime, timedelta

class Wish():
    def __init__(self, academic, date, start_hour, end_hour):
        self.academic = academic
        self.date = date
        self.start_hour = start_hour
        self.end_hour = end_hour

    def check_availability(self, wish_date, wish_start_time, wish_end_time):
        # Convert wish and stored start/end times to timestamp
        # start_datetime_str = f"{self.date} {self.start_hour}"
        # end_datetime_str = f"{self.date} {self.end_hour}"
        # start_timestamp = int(datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S").timestamp())
        # end_timestamp = int(datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M").timestamp())
        #
        # wish_start_timestamp = int(datetime.strptime(f"{wish_date} {wish_start_time}", "%Y-%m-%d %H:%M").timestamp())
        # wish_end_timestamp = int(datetime.strptime(f"{wish_date} {wish_end_time}", "%Y-%m-%d %H:%M").timestamp())

        # Check if the wish is available
        # if (start_timestamp is None or end_timestamp is None or
        #         not (start_timestamp <= wish_start_timestamp <= end_timestamp) or
        #         not (wish_end_timestamp <= start_timestamp or wish_start_timestamp >= end_timestamp)):
        #     return False  # The wish is available
        # else:
        #     return True  # The wish is not available
        return False

class Academic():
    def __init__(self, name):
        self.name = name


import random

import random


import copy
import random
import matplotlib.pyplot as plt

class Population:
    def __init__(self, pop_size, slots, defenses, wish_list):
        self.defense_terms_pop = [DefensesTermsList(self.generate_random_sequence(slots, defenses)) for _ in range(pop_size)]
        self.slots = slots
        self.wish_list = wish_list

    def generate_random_sequence(self, slots, defenses):
        return [DefenseTerm(random.choice(defenses.defense_list), random.choice(slots.slots)) for _ in
                range(min(len(defenses.defense_list), len(slots.slots)))]

    def mutate(self, defense_terms_list, slots):
        mutated_sequence = copy.deepcopy(defense_terms_list)
        index1, index2 = random.sample(range(len(mutated_sequence)), 2)

        # Swap slots in the mutated sequence
        mutated_sequence[index1].slot = random.choice(slots.slots)
        mutated_sequence[index2].slot = random.choice(slots.slots)

        return mutated_sequence

    def cross(self, parent1, parent2, slots):
        crossover_point = random.randint(1, len(parent1.defenses_terms) - 1)
        child1 = DefensesTermsList(parent1.defenses_terms[:crossover_point] + parent2.defenses_terms[crossover_point:])
        child2 = DefensesTermsList(parent2.defenses_terms[:crossover_point] + parent1.defenses_terms[crossover_point:])

        # Swap slots in the crossover children
        for i in range(len(child1.defenses_terms)):
            child1.defenses_terms[i].slot = random.choice(slots.slots)
            child2.defenses_terms[i].slot = random.choice(slots.slots)

        return child1, child2

    def calculate_fitness(self):
        fitness = [individual.fitness(self.wish_list) for individual in self.defense_terms_pop]
        return fitness

    def evolutionary_method(self, generations, top_percentage=0.5, elite_percentage=0.1, mutation_rate=0.7):
        best_individuals_over_time = []

        for generation in range(generations):
            fitness_scores = self.calculate_fitness()

            sorted_indices = sorted(range(len(fitness_scores)), key=lambda k: fitness_scores[k], reverse=True)

            top_indices = sorted_indices[:int(top_percentage * len(fitness_scores))]
            elite_indices = sorted_indices[:int(elite_percentage * len(fitness_scores))]

            elite_individuals = [copy.deepcopy(self.defense_terms_pop[i]) for i in elite_indices]

            print(f"Generation {generation + 1}: Best Fitness = {fitness_scores[sorted_indices[0]]}, Population size = {len(self.defense_terms_pop)}")

            new_population = []
            new_population.extend(elite_individuals)

            for _ in range(len(self.defense_terms_pop) - len(elite_individuals)):
                if random.random() < 1 - mutation_rate:
                    parent1 = random.choice(elite_individuals)
                    parent2 = random.choice(elite_individuals)
                    child1, child2 = self.cross(parent1, parent2, self.slots)
                    new_population.extend([copy.deepcopy(child1), copy.deepcopy(child2)])
                else:
                    mutated_individual = self.mutate(random.choice(elite_individuals).defenses_terms, self.slots)
                    new_population.append(DefensesTermsList(mutated_individual))

            self.defense_terms_pop = new_population[:len(self.defense_terms_pop)]

            best_individuals_over_time.append(max(self.defense_terms_pop, key=lambda x: x.fitness(self.wish_list)))

        best_individual_index = fitness_scores.index(max(fitness_scores))
        best_individual = self.defense_terms_pop[best_individual_index]
        print(f"Evolution finished. Best Fitness = {fitness_scores[best_individual_index]}")

        global_best_individual = max(best_individuals_over_time, key=lambda x: x.fitness(self.wish_list))

        self.plot_evolution(best_individuals_over_time)

        return global_best_individual

    def plot_evolution(self, best_individuals_over_time):
        fitness_values = [individual.fitness(self.wish_list) for individual in best_individuals_over_time]

        plt.plot(range(1, len(best_individuals_over_time) + 1), fitness_values)
        plt.title('Evolution of Fitness Over Generations')
        plt.xlabel('Generation')
        plt.ylabel('Fitness')
        plt.show()
